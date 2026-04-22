"""
Output Formatter - Writes campaign ideas as JSON, markdown, and Excel spreadsheet.
"""

import json
from datetime import datetime
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import CampaignIdea
from config import OPENING_CATEGORIES, ISSUE_DOMAINS, DATA_DIR


def write_json(ideas: list[CampaignIdea], output_path: str = None) -> str:
    """Write ideas to JSON file. Returns the file path."""
    if output_path is None:
        DATA_DIR.mkdir(exist_ok=True)
        output_path = str(DATA_DIR / "ideas.json")

    data = {
        "generated_at": datetime.now().isoformat(),
        "total_ideas": len(ideas),
        "ideas": [i.to_dict() for i in ideas],
    }

    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)

    return output_path


def write_markdown(ideas: list[CampaignIdea], output_path: str = None) -> str:
    """Write ideas as human-readable markdown. Returns the file path."""
    if output_path is None:
        DATA_DIR.mkdir(exist_ok=True)
        output_path = str(DATA_DIR / "ideas.md")

    scored = [i for i in ideas if not i.is_watch_list]
    watch_list = [i for i in ideas if i.is_watch_list]

    lines = []
    lines.append("# Campaign Ideas")
    lines.append(f"\n*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*")
    lines.append(f"\n**Total: {len(ideas)} | Scored: {len(scored)} | Watch list: {len(watch_list)}**")

    # Summary
    lines.append("\n---")
    lines.append("\n## Summary")

    lines.append("\n### By Score Range (scored)")
    score_bands = [("Exceptional (≥3.5)", 3.5, 4.01), ("Strong (≥2.5)", 2.5, 3.5), ("Solid (≥1.5)", 1.5, 2.5), ("Low (<1.5)", 0, 1.5)]
    for label, lo, hi in score_bands:
        count = sum(1 for i in scored if lo <= i.weighted_score < hi)
        if count:
            lines.append(f"- **{label}**: {count}")

    lines.append("\n### By Issue Domain")
    domain_counts = Counter(i.issue_domain for i in ideas)
    for domain in ISSUE_DOMAINS:
        count = domain_counts.get(domain, 0)
        if count > 0:
            lines.append(f"- **{domain}**: {count}")
    # Catch domains not in the standard list
    for domain, count in domain_counts.items():
        if domain and domain not in ISSUE_DOMAINS:
            lines.append(f"- **{domain}**: {count}")

    # Scored ideas ranked by weighted score
    lines.append("\n---")
    lines.append("\n## Scored Campaign Ideas")

    scored_ranked = sorted(scored, key=lambda i: i.weighted_score, reverse=True)
    for rank, idea in enumerate(scored_ranked, 1):
        lines.append(f"\n### {rank}. [{idea.weighted_score:.2f}] {idea.headline}")
        lines.append(f"\n**News hook**: {idea.news_hook}")
        lines.append(f"\n| | |")
        lines.append(f"|---|---|")
        lines.append(f"| **Target** | {idea.target} |")
        lines.append(f"| **Ask** | {idea.ask} |")
        lines.append(f"| **Constituency** | {idea.constituency} |")
        lines.append(f"| **Theory of leverage** | {idea.theory_of_leverage} |")
        lines.append(f"| **Where** | {idea.where} |")
        lines.append(f"| **Time sensitivity** | {idea.time_sensitivity} |")
        lines.append(f"| **Issue domain** | {idea.issue_domain} |")
        lines.append(f"| **Category** | {idea.category} |")

        lines.append(f"\n**Gates**: target={idea.gate_named_target} | ask={idea.gate_binary_ask} | window={idea.gate_time_window}")
        lines.append(f"**Scores**: choir={idea.score_beyond_choir} | pressure={idea.score_pressure_point} | anti-auth={idea.score_anti_authoritarian} | replication={idea.score_replication} | winnability={idea.score_winnability} | energy={idea.score_energy_potential} | non-comply={idea.score_non_compliance}")

        if idea.score_rationale:
            lines.append(f"\n**Rationale**: {idea.score_rationale}")
        if idea.critique_notes:
            lines.append(f"\n**Critique**: {idea.critique_notes}")
        if idea.ai_leverage:
            lines.append(f"\n**AI leverage**: {idea.ai_leverage}")
        if idea.existing_coverage:
            score_label = (
                f" [{idea.coverage_score} — "
                + ["saturated", "crowded", "gap", "wide open"][idea.coverage_score]
                + "]"
                if idea.coverage_score is not None and 0 <= idea.coverage_score <= 3
                else ""
            )
            lines.append(f"\n**Existing coverage**{score_label}: {idea.existing_coverage}")
        if idea.pre_critique_score and idea.pre_critique_score != idea.weighted_score:
            lines.append(f"\n*Score adjusted from {idea.pre_critique_score:.2f} to {idea.weighted_score:.2f} by self-critique*")
        lines.append(f"\n**Source**: [{idea.source_name}]({idea.source_url})")
        if idea.additional_sources:
            for src in idea.additional_sources:
                lines.append(f"- Also: [{src}]({src})")

    # Watch List
    if watch_list:
        lines.append("\n---")
        lines.append(f"\n## Watch List ({len(watch_list)})")
        lines.append("\n*Failed one or more gates but may be worth monitoring.*")

        for rank, idea in enumerate(watch_list, 1):
            lines.append(f"\n**{rank}. {idea.headline}**")
            lines.append(f"- **News hook**: {idea.news_hook}")
            lines.append(f"- **Gates**: target={idea.gate_named_target} | ask={idea.gate_binary_ask} | window={idea.gate_time_window}")
            if idea.gate_fail_reason:
                lines.append(f"- **Why watch list**: {idea.gate_fail_reason}")
            if idea.watch_list_trigger:
                lines.append(f"- **Could become scored if**: {idea.watch_list_trigger}")
            lines.append(f"- **Source**: [{idea.source_name}]({idea.source_url})")

    lines.append("\n---")
    lines.append(f"\n*End of scan. {len(scored)} scored ideas, {len(watch_list)} watch list.*")

    with open(output_path, 'w') as f:
        f.write("\n".join(lines))

    return output_path


def write_xlsx(ideas: list[CampaignIdea], output_path: str = None) -> str:
    """Write ideas to an Excel spreadsheet. Returns the file path."""
    if output_path is None:
        DATA_DIR.mkdir(exist_ok=True)
        output_path = str(DATA_DIR / "ideas.xlsx")

    wb = Workbook()

    # --- Main sheet ---
    ws = wb.active
    ws.title = "Campaign Ideas"

    headers = [
        "Status", "Score", "Headline",
        "Target", "Ask", "Constituency", "Theory of Leverage",
        "News Hook", "Where", "Time Sensitivity",
        "G:Target", "G:Ask", "G:Window", "Gate Fail",
        "D1:Choir", "D2:Pressure", "D3:AntiAuth", "D4:Repl", "D5:Win", "D6:Energy", "D7:NonComply",
        "Rationale", "Critique Notes", "AI Leverage", "Coverage Score", "Existing Coverage", "Pre-Critique Score",
        "Category", "Issue Domain",
        "Source", "Source URL", "Additional Sources",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    def _score_fill(score: float):
        if score >= 3.5: return PatternFill("solid", fgColor="C6EFCE")   # green
        if score >= 2.5: return PatternFill("solid", fgColor="D9E2F3")   # light blue
        if score >= 1.5: return PatternFill("solid", fgColor="FFF2CC")   # light yellow
        return PatternFill("solid", fgColor="FCE4D6")                     # light orange

    watch_list_fill = PatternFill("solid", fgColor="E0E0E0")
    wrap = Alignment(vertical="top", wrap_text=True)

    for row_idx, idea in enumerate(ideas, 2):
        additional = "; ".join(idea.additional_sources) if idea.additional_sources else ""
        status = "Watch List" if idea.is_watch_list else "Scored"

        values = [
            status,
            idea.weighted_score if not idea.is_watch_list else "",
            idea.headline,
            idea.target,
            idea.ask,
            idea.constituency,
            idea.theory_of_leverage,
            idea.news_hook,
            idea.where,
            idea.time_sensitivity,
            idea.gate_named_target,
            idea.gate_binary_ask,
            idea.gate_time_window,
            idea.gate_fail_reason,
            idea.score_beyond_choir if not idea.is_watch_list else "",
            idea.score_pressure_point if not idea.is_watch_list else "",
            idea.score_anti_authoritarian if not idea.is_watch_list else "",
            idea.score_replication if not idea.is_watch_list else "",
            idea.score_winnability if not idea.is_watch_list else "",
            idea.score_energy_potential if not idea.is_watch_list else "",
            idea.score_non_compliance if not idea.is_watch_list else "",
            idea.score_rationale,
            idea.critique_notes,
            idea.ai_leverage,
            idea.coverage_score if idea.coverage_score is not None else "",
            idea.existing_coverage,
            idea.pre_critique_score if idea.pre_critique_score else "",
            idea.category,
            idea.issue_domain,
            idea.source_name,
            idea.source_url,
            additional,
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = wrap
            cell.border = thin_border

        if idea.is_watch_list:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = watch_list_fill
        else:
            fill = _score_fill(idea.weighted_score)
            for col in range(1, 3):
                ws.cell(row=row_idx, column=col).fill = fill

    # Column widths
    col_widths = {
        1: 12, 2: 16, 3: 10, 4: 50,
        5: 30, 6: 35, 7: 30, 8: 40,
        9: 40, 10: 20, 11: 25,
        12: 25,
        13: 8, 14: 8, 15: 8, 16: 30,
        17: 8, 18: 8, 19: 8, 20: 8, 21: 8, 22: 8, 23: 8,
        24: 40, 25: 40, 26: 40, 27: 50, 28: 12,
        29: 25, 30: 25,
        31: 20, 32: 35, 33: 35,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ideas) + 1}"

    # --- Summary sheet ---
    summary = wb.create_sheet("Summary")
    summary_header = Font(bold=True, size=14)
    summary_sub = Font(bold=True, size=11)

    scored = [i for i in ideas if not i.is_watch_list]
    watch_list = [i for i in ideas if i.is_watch_list]

    summary["A1"] = "Campaign Ideas — Summary"
    summary["A1"].font = summary_header
    summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary["A3"] = f"Total: {len(ideas)} | Scored: {len(scored)} | Watch list: {len(watch_list)}"
    summary["A3"].font = Font(bold=True, size=12)

    row = 5
    summary.cell(row=row, column=1, value="By Score Range (scored)").font = summary_sub
    row += 1
    score_bands = [("Exceptional (≥3.5)", 3.5, 4.01), ("Strong (≥2.5)", 2.5, 3.5), ("Solid (≥1.5)", 1.5, 2.5), ("Low (<1.5)", 0, 1.5)]
    for label, lo, hi in score_bands:
        count = sum(1 for i in scored if lo <= i.weighted_score < hi)
        if count:
            summary.cell(row=row, column=1, value=label)
            summary.cell(row=row, column=2, value=count)
            row += 1

    row += 1
    summary.cell(row=row, column=1, value="By Issue Domain").font = summary_sub
    row += 1
    domain_counts = Counter(i.issue_domain for i in ideas)
    for domain in ISSUE_DOMAINS:
        count = domain_counts.get(domain, 0)
        if count > 0:
            summary.cell(row=row, column=1, value=domain)
            summary.cell(row=row, column=2, value=count)
            row += 1
    for domain, count in domain_counts.items():
        if domain and domain not in ISSUE_DOMAINS:
            summary.cell(row=row, column=1, value=domain)
            summary.cell(row=row, column=2, value=count)
            row += 1

    summary.column_dimensions['A'].width = 35
    summary.column_dimensions['B'].width = 10

    wb.save(output_path)
    return output_path


def print_summary(ideas: list[CampaignIdea]) -> None:
    """Print a summary of generated ideas to stdout."""
    scored = [i for i in ideas if not i.is_watch_list]
    watch_list = [i for i in ideas if i.is_watch_list]

    print(f"\n{'='*60}")
    print(f"  SCAN COMPLETE: {len(ideas)} campaign ideas generated")
    print(f"  Scored: {len(scored)} | Watch list: {len(watch_list)}")
    print(f"{'='*60}")

    print("\n  By Score Range (scored):")
    score_bands = [("Exceptional (≥3.5)", 3.5, 4.01), ("Strong (≥2.5)", 2.5, 3.5), ("Solid (≥1.5)", 1.5, 2.5), ("Low (<1.5)", 0, 1.5)]
    for label, lo, hi in score_bands:
        count = sum(1 for i in scored if lo <= i.weighted_score < hi)
        if count:
            print(f"    {label}: {count}")

    top = sorted(scored, key=lambda i: i.weighted_score, reverse=True)[:10]
    print(f"\n  Top {len(top)} Campaign Ideas:")
    for rank, idea in enumerate(top, 1):
        print(f"    {rank}. [{idea.weighted_score:.2f}] {idea.headline[:75]}")
        print(f"       Target: {idea.target[:60]}")
        print(f"       Ask: {idea.ask[:60]}")
        print(f"       Scores: C={idea.score_beyond_choir} P={idea.score_pressure_point} A={idea.score_anti_authoritarian} R={idea.score_replication} W={idea.score_winnability} E={idea.score_energy_potential} N={idea.score_non_compliance}")
        if idea.critique_notes:
            print(f"       Critique: {idea.critique_notes[:70]}")

    if watch_list:
        print(f"\n  Watch List ({len(watch_list)}):")
        for rank, idea in enumerate(watch_list[:5], 1):
            print(f"    {rank}. {idea.headline[:70]}")
            print(f"       Gates: T={idea.gate_named_target} A={idea.gate_binary_ask} W={idea.gate_time_window}")
        if len(watch_list) > 5:
            print(f"    ... and {len(watch_list) - 5} more")

    print(f"\n{'='*60}")
